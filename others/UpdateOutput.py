import openpyxl
from datetime import datetime
# Function to update data in Excel sheet
def update_excel(stock_name, signal):
    # Load the workbook
    wb = openpyxl.load_workbook("output.xlsx")

    # Select the active sheet
    sheet = wb.active

    # Find the first empty row in the sheet
    row = sheet.max_row + 1

    # Update the data in the sheet
    sheet.cell(row=row, column=1).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.cell(row=row, column=2).value = stock_name
    sheet.cell(row=row, column=3).value = signal

    # Save the workbook
    wb.save("output.xlsx")